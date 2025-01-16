from flask import Flask, render_template, request, redirect, url_for, send_file
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import PieChart, Reference

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/baocaoDHSC', methods=['GET', 'POST'])
def baocaoDHSC_route():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            try:
                excel_file = io.BytesIO(file.read())
                output_excel_file = baocaoDHSC(excel_file)
                return send_file(output_excel_file, download_name='thongke_chitiet.xlsx', as_attachment=True)
            except Exception as e:
                return f"Có lỗi xảy ra trong quá trình xử lý: {str(e)}"
    return render_template('baocaoDHSC.html')

@app.route('/baocaoPTTB', methods=['GET', 'POST'])
def baocaoPTTB_route():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            try:
                excel_file = io.BytesIO(file.read())
                output_excel_file = baocaoPTTB(excel_file)
                return send_file(output_excel_file, download_name='baocao_pttb_processed.xlsx', as_attachment=True)
            except Exception as e:
                return f"Có lỗi xảy ra trong quá trình xử lý: {str(e)}"
    return render_template('baocaoPTTB.html')

def baocaoDHSC(excel_file):
    # Đọc file Excel, header ở dòng thứ 2 (index 1)
    df = pd.read_excel(excel_file, header=1)

    # Giá trị DOIVT cần lọc
    allowed_doivt = ['Thạch Thất', 'Sơn Tây', 'Ba Vì', 'Phúc Thọ', 'Đan Phượng']

    # Lọc dữ liệu theo DOIVT
    filtered_df = df[df['DOIVT'].isin(allowed_doivt)].copy()

    # Tạo workbook và writer để ghi dữ liệu vào nhiều sheet
    workbook = Workbook()

    # Sheet ThongKeChung
    sheet_thongke_chung = workbook.create_sheet("ThongKeChung")
    thongke_chung = filtered_df.groupby(['DOIVT', 'NHOMVT']).size().reset_index(name='SoLuong')
    thongke_chung = thongke_chung.sort_values(by='DOIVT')
    for r_idx, row in enumerate(dataframe_to_rows(thongke_chung, header=True, index=False)):
        sheet_thongke_chung.append(row)

    # Sheet ThongKeTheoDOIVT
    for dov in allowed_doivt:
        sheet_dov = workbook.create_sheet(dov)
        thongke_dov = filtered_df[filtered_df['DOIVT'] == dov].groupby('NHOMVT').size().reset_index(name='SoLuong')
        for r_idx, row in enumerate(dataframe_to_rows(thongke_dov, header=True, index=False)):
            sheet_dov.append(row)

        # Vẽ biểu đồ cho từng sheet DOIVT
        chart = BarChart()
        chart.title = f'Thống kê NHOMVT tại {dov}'
        chart.x_axis.title = 'NHOMVT'
        chart.y_axis.title = 'Số lượng'
        data = Reference(sheet_dov, min_col=2, min_row=1, max_col=2, max_row=sheet_dov.max_row)
        chart.add_data(data)
        cats = Reference(sheet_dov, min_col=1, min_row=2, max_col=1, max_row=sheet_dov.max_row)
        chart.set_categories(cats)
        sheet_dov.add_chart(chart, "E2")

    # Sheet ThongKeDOIVTTong
    sheet_doivt_tong = workbook.create_sheet("ThongKeDOIVTTong")
    thongke_doivt_tong = filtered_df['DOIVT'].value_counts().reset_index()
    thongke_doivt_tong.columns = ['DOIVT', 'SoLuong']
    for r_idx, row in enumerate(dataframe_to_rows(thongke_doivt_tong, header=True, index=False)):
        sheet_doivt_tong.append(row)

    # Vẽ biểu đồ cho sheet ThongKeDOIVTTong
    chart_tong_dov = BarChart()
    chart_tong_dov.title = 'Tổng số lượng theo DOIVT'
    chart_tong_dov.x_axis.title = 'DOIVT'
    chart_tong_dov.y_axis.title = 'Số lượng'
    data_tong_dov = Reference(sheet_doivt_tong, min_col=2, min_row=1, max_col=2, max_row=sheet_doivt_tong.max_row)
    chart_tong_dov.add_data(data_tong_dov)
    cats_tong_dov = Reference(sheet_doivt_tong, min_col=1, min_row=2, max_col=1, max_row=sheet_doivt_tong.max_row)
    chart_tong_dov.set_categories(cats_tong_dov)
    sheet_doivt_tong.add_chart(chart_tong_dov, "E2")

    # Sheet Top10NHOMVT
    sheet_top10_nhomvt = workbook.create_sheet("Top10NHOMVT")
    top10_nhomvt = filtered_df['NHOMVT'].value_counts().nlargest(10).reset_index()
    top10_nhomvt.columns = ['NHOMVT', 'SoLuong']
    for r_idx, row in enumerate(dataframe_to_rows(top10_nhomvt, header=True, index=False)):
        sheet_top10_nhomvt.append(row)

    # Vẽ biểu đồ cho sheet Top10NHOMVT
    chart_top10 = BarChart()
    chart_top10.title = 'Top 10 NHOMVT'
    chart_top10.x_axis.title = 'NHOMVT'
    chart_top10.y_axis.title = 'Số lượng'
    data_top10 = Reference(sheet_top10_nhomvt, min_col=2, min_row=1, max_col=2, max_row=sheet_top10_nhomvt.max_row)
    chart_top10.add_data(data_top10)
    cats_top10 = Reference(sheet_top10_nhomvt, min_col=1, min_row=2, max_col=1, max_row=sheet_top10_nhomvt.max_row)
    chart_top10.set_categories(cats_top10)
    sheet_top10_nhomvt.add_chart(chart_top10, "E2")

    # Lưu workbook vào bộ nhớ
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output




def baocaoPTTB(excel_file):
    # Đọc file Excel, header ở dòng thứ 2 (index 1)
    df = pd.read_excel(excel_file, header=1)

    # Các giá trị DOIVT_KV cần lọc
    allowed_doivt_kv = ['Thạch Thất', 'Ba Vì', 'Đan Phượng', 'Sơn Tây', 'Phúc Thọ']

    # Lọc dữ liệu theo DOIVT_KV
    df_filtered = df[df['DOIVT_KV'].isin(allowed_doivt_kv)].copy()

    # Tạo workbook
    workbook = Workbook()

    # Tạo sheet tổng hợp
    sheet_tonghop = workbook.create_sheet(title="TongHop")

    # Lấy danh sách các loại hình TB cần thống kê
    loaihinh_tb_columns = [
        'Điện thoại cố định', 'Megawan quang FE', 'Fiber', 'Thuê bao SIP', 'MetroNet GE',
        'Cáp quang trắng', 'VNPT Family Safe', 'MetroNet FE', 'Metronet_POP', 'MyTV',
        'Wifi Mesh', 'Indoor Camera PT', 'Home Cloud camera'
    ]

    # Tạo bảng thống kê tổng hợp ban đầu
    thongke_pttb = df_filtered.groupby('DOIVT_KV').size().reset_index(name='số máy tồn')
    thongke_pttb = thongke_pttb.sort_values(by='DOIVT_KV').reset_index(drop=True)
    thongke_pttb.insert(0, 'STT', range(1, len(thongke_pttb) + 1))

    # Thêm các cột loại hình TB vào bảng thống kê
    for col in loaihinh_tb_columns:
        thongke_pttb[col] = 0  # Khởi tạo cột với giá trị 0

    # Điền dữ liệu vào các cột loại hình TB
    for index, row in thongke_pttb.iterrows():
        doivt_kv = row['DOIVT_KV']
        df_doivt = df_filtered[df_filtered['DOIVT_KV'] == doivt_kv]
        for col in loaihinh_tb_columns:
            count = len(df_doivt[df_doivt['LOAIHINH_TB'] == col])
            thongke_pttb.loc[index, col] = count

    # Ghi dữ liệu vào sheet TongHop
    for r_idx, row in enumerate(dataframe_to_rows(thongke_pttb, header=True, index=False)):
        sheet_tonghop.append(row)

    # Vẽ biểu đồ tròn
    pie_chart = PieChart()
    pie_chart.title = 'Tỷ lệ số máy tồn theo DOIVT_KV'
    labels = Reference(sheet_tonghop, min_col=2, min_row=2, max_row=sheet_tonghop.max_row)
    data = Reference(sheet_tonghop, min_col=3, min_row=1, max_row=sheet_tonghop.max_row)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    sheet_tonghop.add_chart(pie_chart, "F1")

    # Tạo các sheet chi tiết
    for doivt in allowed_doivt_kv:
        df_doivt = df[df['DOIVT_KV'] == doivt].copy()
        if not df_doivt.empty:
            # Chuyển cột DOIVT_KV lên đầu
            doivt_kv_col = df_doivt.pop('DOIVT_KV')
            df_doivt.insert(0, 'DOIVT_KV', doivt_kv_col)

            sheet_chi_tiet = workbook.create_sheet(title=doivt)
            for r_idx, row in enumerate(dataframe_to_rows(df_doivt, header=True, index=False)):
                sheet_chi_tiet.append(row)

    # Tạo bảng mới cho Thạch Thất
    df_thachthat = df[df['DOIVT_KV'] == "Thạch Thất"].copy()
    thongke_tenkv_thachthat = df_thachthat.groupby('TEN_KV').size().reset_index(name='Số lượng')
    thongke_tenkv_thachthat.insert(0, 'STT', range(1, len(thongke_tenkv_thachthat) + 1))

    # Tạo sheet và ghi dữ liệu cho Thạch Thất
    sheet_thachthat_tenkv = workbook.create_sheet(title="ThachThat_TENKV")
    for r_idx, row in enumerate(dataframe_to_rows(thongke_tenkv_thachthat, header=True, index=False)):
        sheet_thachthat_tenkv.append(row)

    # Tạo bảng mới cho Ba Vì
    df_bavi = df[df['DOIVT_KV'] == "Ba Vì"].copy()
    thongke_tenkv_bavi = df_bavi.groupby('TEN_KV').size().reset_index(name='Số lượng')
    thongke_tenkv_bavi.insert(0, 'STT', range(1, len(thongke_tenkv_bavi) + 1))

    # Tạo sheet và ghi dữ liệu cho Ba Vì
    sheet_bavi_tenkv = workbook.create_sheet(title="BaVi_TENKV")
    for r_idx, row in enumerate(dataframe_to_rows(thongke_tenkv_bavi, header=True, index=False)):
        sheet_bavi_tenkv.append(row)

    # Tạo bảng mới cho Phúc Thọ
    df_phuctho = df[df['DOIVT_KV'] == "Phúc Thọ"].copy()
    thongke_tenkv_phuctho = df_phuctho.groupby('TEN_KV').size().reset_index(name='Số lượng')
    thongke_tenkv_phuctho.insert(0, 'STT', range(1, len(thongke_tenkv_phuctho) + 1))

    # Tạo sheet và ghi dữ liệu cho Phúc Thọ
    sheet_phuctho_tenkv = workbook.create_sheet(title="PhucTho_TENKV")
    for r_idx, row in enumerate(dataframe_to_rows(thongke_tenkv_phuctho, header=True, index=False)):
        sheet_phuctho_tenkv.append(row)

    # Tạo bảng mới cho Sơn Tây
    df_sontay = df[df['DOIVT_KV'] == "Sơn Tây"].copy()
    thongke_tenkv_sontay = df_sontay.groupby('TEN_KV').size().reset_index(name='Số lượng')
    thongke_tenkv_sontay.insert(0, 'STT', range(1, len(thongke_tenkv_sontay) + 1))

    # Tạo sheet và ghi dữ liệu cho Sơn Tây
    sheet_sontay_tenkv = workbook.create_sheet(title="SonTay_TENKV")
    for r_idx, row in enumerate(dataframe_to_rows(thongke_tenkv_sontay, header=True, index=False)):
        sheet_sontay_tenkv.append(row)

    # Tạo bảng mới cho Đan Phượng
    df_danphuong = df[df['DOIVT_KV'] == "Đan Phượng"].copy()
    thongke_tenkv_danphuong = df_danphuong.groupby('TEN_KV').size().reset_index(name='Số lượng')
    thongke_tenkv_danphuong.insert(0, 'STT', range(1, len(thongke_tenkv_danphuong) + 1))

    # Tạo sheet và ghi dữ liệu cho Đan Phượng
    sheet_danphuong_tenkv = workbook.create_sheet(title="DanPhuong_TENKV")
    for r_idx, row in enumerate(dataframe_to_rows(thongke_tenkv_danphuong, header=True, index=False)):
        sheet_danphuong_tenkv.append(row)

    # Loại bỏ sheet mặc định nếu nó trống
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Lưu workbook vào bộ nhớ
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output




if __name__ == '__main__':
    app.run(debug=True)